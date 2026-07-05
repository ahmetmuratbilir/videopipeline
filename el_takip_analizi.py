import cv2
import csv
import numpy as np
from fsm import MOSTTracker, point_in_polygon

from config_manager import config as APP_CONFIG

# ── Ayarlanabilir Sabitler (config_manager'dan yüklenir) ──────────────────────
EL_KAYIP_SURE        = APP_CONFIG["EL_KAYIP_SURE"]
EL_GRACE_SURE        = APP_CONFIG["EL_GRACE_SURE"]
EL_CONFIDENCE_ESIGI  = APP_CONFIG["EL_CONFIDENCE_ESIGI"]
ERGO_RISK_YUKSEK     = APP_CONFIG["ERGO_RISK_YUKSEK"]
ERGO_RISK_DIKKAT     = APP_CONFIG["ERGO_RISK_DIKKAT"]
GOVDE_RISK_YUKSEK    = APP_CONFIG["GOVDE_RISK_YUKSEK"]
GOVDE_RISK_DIKKAT    = APP_CONFIG["GOVDE_RISK_DIKKAT"]
FLASH_KARE_SAYISI    = APP_CONFIG["FLASH_KARE_SAYISI"]
HUD_SERIT_GENISLIK   = APP_CONFIG["HUD_SERIT_GENISLIK"]
GEMINI_MODEL_NAME    = APP_CONFIG["GEMINI_MODEL_NAME"]

# ── MediaPipe ─────────────────────────────────────────────────────────────────
try:
    import mediapipe as mp
    _mp_hands = mp.solutions.hands
    _mp_pose  = mp.solutions.pose
    _mp_drawing = mp.solutions.drawing_utils
    _mp_drawing_styles = mp.solutions.drawing_styles

    eller_dedektoru = _mp_hands.Hands(
        max_num_hands=2,
        min_detection_confidence=EL_CONFIDENCE_ESIGI,
        min_tracking_confidence=EL_CONFIDENCE_ESIGI
    )
    pose_dedektoru = _mp_pose.Pose(
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5,
        model_complexity=1   # 0=hızlı, 1=dengeli, 2=en doğru
    )
    MEDIAPIPE_HAZIR = True
except Exception as e:
    print(f"[UYARI] MediaPipe yüklenemedi, simülasyon moduna geçiliyor: {e}")
    MEDIAPIPE_HAZIR = False

# ── xlsx isteğe bağlı ─────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_HAZIR = True
except ImportError:
    XLSX_HAZIR = False


# ── Yardımcı Fonksiyonlar ─────────────────────────────────────────────────────

def alan_poligonunu_al(veri):
    """Poligon varsa onu, yoksa bounding box'tan 4 köşeli poligon üretir."""
    if "polygon" in veri and veri["polygon"]:
        return veri["polygon"]
    x, y, w, h = veri["x"], veri["y"], veri["w"], veri["h"]
    return [[x, y], [x + w, y], [x + w, y + h], [x, y + h]]


def calculate_angle(a, b, c):
    """Üç 2B nokta arasındaki açıyı derece cinsinden hesaplar.
    b noktası açının köşesidir (dirsek). optiMOST / unified_analyzer.py ile aynı mantık."""
    a = np.array(a, dtype=np.float32)
    b = np.array(b, dtype=np.float32)
    c = np.array(c, dtype=np.float32)
    ba, bc = a - b, c - b
    cosine = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-6)
    return float(np.degrees(np.arccos(np.clip(cosine, -1.0, 1.0))))


def calculate_trunk_angle(hip, shoulder):
    """
    Kalça-omuz hattının dikey eksenden (yukarı vektör) sapma açısını hesaplar.
    0° = tam dikey (ideal duruş), 90° = tam yatay (öne eğilme).
    """
    hip = np.array(hip, dtype=np.float32)
    shoulder = np.array(shoulder, dtype=np.float32)
    trunk_vector = shoulder - hip
    norm = np.linalg.norm(trunk_vector)
    if norm < 1e-6:
        return None
        
    vertical_vector = np.array([0, -1], dtype=np.float32)  # Görüntü koordinatında "yukarı"
    
    cosine_angle = np.dot(trunk_vector, vertical_vector) / (norm * np.linalg.norm(vertical_vector))
    cosine_angle = np.clip(cosine_angle, -1.0, 1.0)
    return float(np.degrees(np.arccos(cosine_angle)))


def aci_rengini_al(aci, govde=False):
    """Ergonomi risk rengi (sabitlerden): ≥YUKSEK kırmızı, ≥DIKKAT sarı, altı yeşil."""
    yuksek = GOVDE_RISK_YUKSEK if govde else ERGO_RISK_YUKSEK
    dikkat = GOVDE_RISK_DIKKAT if govde else ERGO_RISK_DIKKAT
    if aci >= yuksek:
        return (0, 0, 255)
    elif aci >= dikkat:
        return (0, 180, 255)
    return (0, 200, 0)


def aci_etiketi(aci, govde=False):
    """Renk körü dostu metin etiketi: renk + sembol + metin."""
    yuksek = GOVDE_RISK_YUKSEK if govde else ERGO_RISK_YUKSEK
    dikkat = GOVDE_RISK_DIKKAT if govde else ERGO_RISK_DIKKAT
    if aci >= yuksek:
        return "X Yuksek"
    elif aci >= dikkat:
        return "! Dikkat"
    return "  Normal"


def alanlardan_fsm_config_uret(tanimlanan_alanlar):
    """self.alanlar → MOSTTracker workspace_config. Alet Alanları sıraya göre recipe'ye girer."""
    stations, recipe = [], []
    for alan_adi, veri in tanimlanan_alanlar.items():
        polygon = alan_poligonunu_al(veri)
        if veri["tip"] == "Calisma Alanı":
            fsm_adi = "Assembly Area"
        else:
            fsm_adi = alan_adi
            recipe.append(fsm_adi)
        stations.append({"name": fsm_adi, "polygon": polygon})
    recipe.append("Assembly Area")
    return {
        "stations": stations,
        "recipe": recipe,
        "grasp": {
            "pinch_threshold": 0.28,
            "release_threshold": 0.40,
            "velocity_threshold": 8.0,
            "grasp_confirm_time": 0.15,
            "release_confirm_time": 0.10
        }
    }


def el_landmarklarini_sozluge_cevir(el_noktalari):
    return [{"x": lm.x, "y": lm.y} for lm in el_noktalari.landmark]


def el_durumu_guncelle(ctx, tum_el_noktalari, kare_zamani, kare_suresi, yukseklik, genislik):
    """
    Grace period state machine — sol ve sağ el bağımsız işlenir.
    Zaman bazı: video timestamp kullanılır (kare_zamani), time.time() değil.
    """
    yeni_pozisyonlar = []
    for el in tum_el_noktalari:
        bilek = el.landmark[0]
        yeni_pozisyonlar.append((int(bilek.x * genislik), int(bilek.y * yukseklik)))

    def en_yakin_pozisyon(hedef_poz, adaylar, kullanildi):
        if hedef_poz is None or not adaylar:
            return None, None
        en_kisa, en_i = float('inf'), None
        for i, poz in enumerate(adaylar):
            if i in kullanildi:
                continue
            d = ((poz[0]-hedef_poz[0])**2 + (poz[1]-hedef_poz[1])**2)**0.5
            if d < en_kisa:
                en_kisa, en_i = d, i
        if en_i is not None and en_kisa < 200:  # 200px eşiği aşarsa yeni el
            return adaylar[en_i], en_i
        return None, None

    kullanildi_indeksler = set()
    el_yuvalari = [ctx.sol_el, ctx.sag_el]

    for el_durumu in el_yuvalari:
        eslesik_poz, idx = en_yakin_pozisyon(el_durumu.son_pozisyon, yeni_pozisyonlar, kullanildi_indeksler)
        if eslesik_poz is not None:
            kullanildi_indeksler.add(idx)
            el_durumu.ilk_goruldu = True
            el_durumu.aktif = True
            el_durumu.grace_aktif = False
            el_durumu.dondurulmus = False
            el_durumu.kayip_baslangic_zamani = 0.0
            el_durumu.son_pozisyon = eslesik_poz
        else:
            if el_durumu.aktif:
                el_durumu.aktif = False
                el_durumu.kayip_baslangic_zamani = kare_zamani

            if not el_durumu.aktif and el_durumu.ilk_goruldu:
                kayip_sure = kare_zamani - el_durumu.kayip_baslangic_zamani
                if kayip_sure >= EL_KAYIP_SURE and not el_durumu.grace_aktif:
                    el_durumu.grace_aktif = True
                    el_durumu.grace_baslangic_zamani = kare_zamani
                    el_durumu.dondurulmus = True
                    ctx.el_kayip_olay_sayisi += 1

                if el_durumu.grace_aktif:
                    grace_sure = kare_zamani - el_durumu.grace_baslangic_zamani
                    ctx.el_kayip_toplam_sure += kare_suresi
                    ctx.hand_state.hareket_sureleri["El Kaybi"] += kare_suresi
                    if grace_sure >= EL_GRACE_SURE:
                        el_durumu.grace_aktif = False
                        el_durumu.dondurulmus = False
                        ctx.fsm_events.append({
                            "type": "EL_KAYBI",
                            "timestamp": round(kare_zamani, 3),
                            "sure": round(grace_sure, 3)
                        })

    # Eşleştirilemeyen yeni pozisyonları boş yuvaya ata
    for i, poz in enumerate(yeni_pozisyonlar):
        if i in kullanildi_indeksler:
            continue
        for el_durumu in el_yuvalari:
            if not el_durumu.aktif:
                el_durumu.ilk_goruldu = True
                el_durumu.aktif = True
                el_durumu.grace_aktif = False
                el_durumu.dondurulmus = False
                el_durumu.kayip_baslangic_zamani = 0.0
                el_durumu.son_pozisyon = poz
                break


def cycle_tracker_guncelle(ctx, tanimlanan_alanlar, tum_el_noktalari, kare_zamani, yukseklik, genislik):
    """
    Çevrim süresini ve ürün sayısını takip eder.
    İki el senaryosunda OR mantığı: herhangi bir el alanda mı diye bakılır.
    """
    ct = ctx.cycle_tracker

    # Alan poligonlarını tiple bul
    yesil_poligonlar = [
        alan_poligonunu_al(v) for v in tanimlanan_alanlar.values()
        if v["tip"] == "Calisma Alanı"
    ]
    bitis_poligonlar = [
        alan_poligonunu_al(v) for v in tanimlanan_alanlar.values()
        if v["tip"] == "Bitis Alanı"
    ]

    if not yesil_poligonlar and not bitis_poligonlar:
        return

    # Sadece aktif (dondurulmamış) el konumlarını sayalım
    el_merkezleri = []
    if ctx.sol_el.aktif and not ctx.sol_el.dondurulmus and ctx.sol_el.son_pozisyon:
        el_merkezleri.append(ctx.sol_el.son_pozisyon)
    if ctx.sag_el.aktif and not ctx.sag_el.dondurulmus and ctx.sag_el.son_pozisyon:
        el_merkezleri.append(ctx.sag_el.son_pozisyon)

    # OR mantığı: herhangi bir el alanda mı?
    el_yesilde = any(
        any(point_in_polygon(merkez, poly) for poly in yesil_poligonlar)
        for merkez in el_merkezleri
    ) if el_merkezleri and yesil_poligonlar else False

    el_bitis_alaninda = any(
        any(point_in_polygon(merkez, poly) for poly in bitis_poligonlar)
        for merkez in el_merkezleri
    ) if el_merkezleri and bitis_poligonlar else False

    # 1. İlk yeşil giriş koru
    if el_yesilde and not ct.ilk_yesil_giris_yapildi:
        ct.ilk_yesil_giris_yapildi = True

    # 2. Yeşil'den çıkış → döngü başlar
    if not el_yesilde and ct.yesil_alanda and ct.ilk_yesil_giris_yapildi:
        ct.dongu_baslangic_zamani = kare_zamani
        ct.son_bitis_sonrasi_yesile_gidildi = True

    # 3. Bitiş alanına yeni giriş
    if el_bitis_alaninda and not ct.bitis_alaninda:
        if ct.son_bitis_sonrasi_yesile_gidildi:
            ct.bitis_giris_zamani = kare_zamani
            ct.bu_giriste_sayildi = False

    # 4. Debounce + sayım
    if (el_bitis_alaninda
            and ct.ilk_yesil_giris_yapildi
            and ct.son_bitis_sonrasi_yesile_gidildi
            and not ct.bu_giriste_sayildi):
        bitis_kalma = kare_zamani - ct.bitis_giris_zamani
        if bitis_kalma >= ct.bitis_debounce_sure:
            sure = kare_zamani - ct.dongu_baslangic_zamani
            if 0.5 < sure <= ct.max_dongu_sure:
                ct.dongu_sureleri.append(round(sure, 3))
                ct.urun_sayisi += 1
                ct.bu_giriste_sayildi = True
                ct.son_bitis_sonrasi_yesile_gidildi = False
                
                # Flash animasyon tetikle
                ctx.flash_kare_sayaci = FLASH_KARE_SAYISI
                ctx.flash_alan_tipi = "Bitis"
            elif sure > ct.max_dongu_sure:
                ct.atlanan_dongu_sureleri.append(round(sure, 3))
                ct.bu_giriste_sayildi = True
                print(f"[UYARI] Şüpheli uzun döngü ({sure:.1f}s) atlandı.")

    # 5. Durum güncelleme — EN SONDA
    ct.yesil_alanda = el_yesilde
    ct.bitis_alaninda = el_bitis_alaninda


# ── HUD: Sağ Dikey Şerit (tek panel) ──────────────────────────────────────────

def _yari_saydam_dikdortgen(kare, x1, y1, x2, y2, renk=(0, 0, 0), alfa=0.55):
    overlay = kare.copy()
    cv2.rectangle(overlay, (x1, y1), (x2, y2), renk, -1)
    cv2.addWeighted(overlay, alfa, kare, 1 - alfa, 0, kare)


def _bar(kare, x, y, genislik, oran, on_renk, yukseklik=8):
    cv2.rectangle(kare, (x, y), (x + genislik, y + yukseklik), (40, 40, 40), -1)
    dolu = int(genislik * max(0.0, min(1.0, oran)))
    if dolu > 0:
        cv2.rectangle(kare, (x, y), (x + dolu, y + yukseklik), on_renk, -1)
    cv2.rectangle(kare, (x, y), (x + genislik, y + yukseklik), (80, 80, 80), 1)


def _serit_baslik(kare, x, y, metin, renk=(160, 160, 220)):
    cv2.putText(kare, metin, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.38, renk, 1)


def _serit_satir(kare, x, y, metin, renk=(220, 220, 220)):
    cv2.putText(kare, metin, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.42, renk, 1)


def _ayirici(kare, x1, x2, y, renk=(60, 60, 80)):
    cv2.line(kare, (x1, y), (x2, y), renk, 1)


def flash_ciz(kare, ctx, tanimlanan_alanlar):
    """Alan geçişi flash animasyonu — frame sayacı bazılı (FPS bağımsız)."""
    if ctx.flash_kare_sayaci <= 0:
        return
    ctx.flash_kare_sayaci -= 1
    alfas = 0.35 * (ctx.flash_kare_sayaci / FLASH_KARE_SAYISI)
    renk = (0, 200, 60) if ctx.flash_alan_tipi == "Bitis" else (200, 80, 0)
    for veri in tanimlanan_alanlar.values():
        if ctx.flash_alan_tipi == "Bitis" and veri["tip"] == "Bitis Alanı":
            pts = np.array(alan_poligonunu_al(veri), dtype=np.int32)
            overlay = kare.copy()
            cv2.fillPoly(overlay, [pts], renk)
            cv2.addWeighted(overlay, alfas, kare, 1 - alfas, 0, kare)
        elif ctx.flash_alan_tipi == "Alet" and veri["tip"] == "Alet Alanı":
            pts = np.array(alan_poligonunu_al(veri), dtype=np.int32)
            overlay = kare.copy()
            cv2.fillPoly(overlay, [pts], renk)
            cv2.addWeighted(overlay, alfas, kare, 1 - alfas, 0, kare)


def el_kayip_banner_ciz(kare, ctx, kare_zamani):
    """Grace period aktifse üst ortaya banner çizer. Startup'ta gösterilmez."""
    sol_grace = ctx.sol_el.grace_aktif
    sag_grace = ctx.sag_el.grace_aktif
    if not (sol_grace or sag_grace):
        return
    h, w = kare.shape[:2]
    kalan_sureler = []
    if sol_grace:
        kalan_sureler.append(EL_GRACE_SURE - (kare_zamani - ctx.sol_el.grace_baslangic_zamani))
    if sag_grace:
        kalan_sureler.append(EL_GRACE_SURE - (kare_zamani - ctx.sag_el.grace_baslangic_zamani))
    kalan = max(0.0, max(kalan_sureler))
    metin = f"EL TESPIT EDILEMEDI  {kalan:.1f}s"
    (tw, th), _ = cv2.getTextSize(metin, cv2.FONT_HERSHEY_SIMPLEX, 0.60, 2)
    bx = (w - tw) // 2 - 10
    by = 14
    overlay = kare.copy()
    cv2.rectangle(overlay, (bx - 5, by - 2), (bx + tw + 15, by + th + 6), (0, 0, 160), -1)
    cv2.addWeighted(overlay, 0.75, kare, 0.25, 0, kare)
    cv2.putText(kare, metin, (bx + 5, by + th),
                cv2.FONT_HERSHEY_SIMPLEX, 0.60, (80, 180, 255), 2)


def hud_serit_ciz(kare, ctx, o_anki_durum, fsm_metrikleri, fsm_recipe, fsm_recipe_idx):
    """
    Sağ kenar HUD şeridi — eski iki panelin yerini alır.
    Video alanı şeridin sol tarafında kalır.
    """
    h, w = kare.shape[:2]
    ct = ctx.cycle_tracker
    sol_aci = ctx.ergo_state.son_sol_aci
    sag_aci = ctx.ergo_state.son_sag_aci
    govde_acisi = ctx.ergo_state.son_govde_acisi

    sx = w - HUD_SERIT_GENISLIK   # şeridin sol kenarı
    ex = w - 1                    # şeridin sağ kenarı
    ic = sx + 8                   # iç metin başlangıç
    bar_genislik = HUD_SERIT_GENISLIK - 20

    # Arka plan
    _yari_saydam_dikdortgen(kare, sx, 0, ex, h, renk=(8, 10, 22), alfa=0.75)
    cv2.rectangle(kare, (sx, 0), (ex, h), (60, 60, 120), 1)

    y = 18

    # ─ Ürün Takibi ────────────────────────────────────────────
    _serit_baslik(kare, ic, y, "URUN TAKIBI", (80, 220, 80))
    y += 18
    cv2.putText(kare, f"Urun: {ct.urun_sayisi}", (ic, y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.58, (60, 255, 140), 2)
    y += 20
    if ct.dongu_sureleri:
        ort = ct.cevrim_suresi
        sapma = ct.standart_sapma
        sapma_str = f"+-{sapma:.1f}" if sapma else "N/A"
        _serit_satir(kare, ic, y, f"Cevrim: {ort:.1f}s {sapma_str}", (180, 240, 180))
        y += 16
        son3 = ct.dongu_sureleri[-3:]
        _serit_satir(kare, ic, y, "Son: " + " | ".join(f"{s:.1f}" for s in son3), (130, 190, 130))
    else:
        _serit_satir(kare, ic, y, "Bekleniyor...", (100, 140, 100))
    y += 18
    _ayirici(kare, sx + 4, ex - 4, y)
    y += 10

    # ─ Anlık Durum ────────────────────────────────────────────
    _serit_baslik(kare, ic, y, "AN", (160, 160, 220))
    y += 16
    durum_renkleri = {
        "Montaj / Calisma":              (0, 255, 120),
        "Malzeme Alma / Kavrama":        (0, 200, 255),
        "Malzeme Tasimi (Masaya Dogru)": (255, 200, 0),
        "Kutulara Dogru Uzanma":         (255, 160, 40),
        "Bosta (Bekleme)":               (160, 160, 160),
        "El Kaybi":                      (80, 80, 200),
    }
    d_renk = durum_renkleri.get(o_anki_durum, (200, 200, 200))
    kisa = o_anki_durum.replace(" / ", "/").replace(" (Masaya Dogru)", "").replace(" Dogru", "")
    cv2.putText(kare, kisa, (ic, y), cv2.FONT_HERSHEY_SIMPLEX, 0.44, d_renk, 1)
    y += 14
    _ayirici(kare, sx + 4, ex - 4, y)
    y += 10

    # ─ Durum Geçmişi (tum oturum yüzdeleri) ──────────────────────
    _serit_baslik(kare, ic, y, "GECMIS", (160, 160, 220))
    y += 16
    toplam_sure = sum(ctx.hand_state.hareket_sureleri.values()) or 1.0
    bar_renkleri = {
        "Montaj / Calisma":              (0, 200, 80),
        "Malzeme Alma / Kavrama":        (0, 180, 220),
        "Malzeme Tasimi (Masaya Dogru)": (200, 160, 0),
        "Kutulara Dogru Uzanma":         (200, 120, 30),
        "Bosta (Bekleme)":               (120, 120, 120),
        "El Kaybi":                      (60, 60, 180),
    }
    kisa_isimler = {
        "Montaj / Calisma":              "Montaj",
        "Malzeme Alma / Kavrama":        "Alet Al",
        "Malzeme Tasimi (Masaya Dogru)": "Tasima",
        "Kutulara Dogru Uzanma":         "Uzanma",
        "Bosta (Bekleme)":               "Bosta",
        "El Kaybi":                      "El Kaybi",
    }
    for hareket, sure in ctx.hand_state.hareket_sureleri.items():
        oran = sure / toplam_sure
        renk = bar_renkleri.get(hareket, (120, 120, 120))
        kisa_isim = kisa_isimler.get(hareket, hareket[:8])
        _serit_satir(kare, ic, y, f"{kisa_isim[:7]:<7} {oran*100:4.0f}%", renk)
        y += 12
        _bar(kare, ic, y, bar_genislik, oran, renk)
        y += 12
    _ayirici(kare, sx + 4, ex - 4, y)
    y += 10

    # ─ Ergonomi ───────────────────────────────────────────────
    _serit_baslik(kare, ic, y, "ERGONOMI", (160, 160, 220))
    y += 16

    def ergo_satir(etiket, aci, don, govde=False):
        nonlocal y
        if aci is not None:
            renk = aci_rengini_al(aci, govde)
            etiket_str = "~" if don else " "
            deger_str = f"{etiket_str}{aci:.0f} {aci_etiketi(aci, govde)}"
        else:
            renk = (100, 100, 100)
            deger_str = "---"
        cv2.putText(kare, f"{etiket}: {deger_str}", (ic, y),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, renk, 1)
        y += 14

    sol_don = ctx.sol_el.dondurulmus
    sag_don = ctx.sag_el.dondurulmus
    ergo_satir("Sol", sol_aci, sol_don)
    ergo_satir("Sag", sag_aci, sag_don)
    ergo_satir("Govde", govde_acisi, False, govde=True)
    y += 2
    cv2.putText(kare, "Kamera sabit olmali!", (ic, y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.30, (80, 80, 120), 1)
    y += 12
    _ayirici(kare, sx + 4, ex - 4, y)
    y += 10

    # ─ FSM / MOST Bilgisi ──────────────────────────────────────────
    if fsm_metrikleri:
        _serit_baslik(kare, ic, y, "SIRADAKI ADIM", (160, 160, 220))
        y += 16
        if fsm_recipe and len(fsm_recipe) > 0:
            toplam_r = len(fsm_recipe)
            gunceli = min(fsm_recipe_idx, toplam_r)
            siradaki = fsm_recipe[min(gunceli, toplam_r - 1)]
            _serit_satir(kare, ic, y, f"{gunceli}/{toplam_r}: {siradaki[:14]}", (200, 220, 255))
            y += 12
            _bar(kare, ic, y, bar_genislik, gunceli / toplam_r if toplam_r else 0, (80, 140, 255))
            y += 14
        fsm_state = fsm_metrikleri.get("state", "IDLE")
        fsm_durum_tr = {
            "IDLE": "Bekleme", "REACH": "Uzaniyor", "GRASP": "Kavruyor",
            "MOVE": "Tasiyor", "PLACE": "Birakiyor", "RETURNING_HOME": "Donuyor"
        }
        cycle_sn = fsm_metrikleri.get("cycle_time_sec", 0.0)
        seq_err  = fsm_metrikleri.get("sequence_error", False)
        fsm_renk = (80, 80, 255) if seq_err else (180, 220, 255)
        _serit_satir(kare, ic, y, f"{fsm_durum_tr.get(fsm_state, fsm_state)}  {cycle_sn:.1f}s", fsm_renk)
        y += 14
        if seq_err:
            cv2.putText(kare, "! SIRA HATASI", (ic, y),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0, 60, 255), 2)
            y += 16


# ── Ana Analiz Fonksiyonu ─────────────────────────────────────────────────────

def el_takip_ve_ergonomi_motoru_kare(kare, tanimlanan_alanlar, ctx, kare_suresi, fps, target_size=None):
    yukseklik, genislik, _ = kare.shape

    ctx.kare_sayaci += 1
    kare_sayaci = ctx.kare_sayaci
    kare_zamani = kare_sayaci * kare_suresi

    # 1. Alan Poligonlarını Çizme
    for alan_adi, veri in tanimlanan_alanlar.items():
        if veri["tip"] == "Calisma Alanı":
            renk = (0, 255, 0)
        elif veri["tip"] == "Alet Alanı":
            renk = (0, 0, 255)
        elif veri["tip"] == "Bitis Alanı":
            renk = (255, 100, 0)
        else:
            renk = (128, 128, 128)
        poligon = alan_poligonunu_al(veri)
        pts = np.array(poligon, dtype=np.int32)
        cv2.polylines(kare, [pts], isClosed=True, color=renk, thickness=2)
        etiket_x, etiket_y = int(pts[:, 0].min()), int(pts[:, 1].min())
        cv2.putText(kare,
                    f"{alan_adi}",
                    (etiket_x, etiket_y - 8),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, renk, 1)

    elin_oldugu_aktif_alan = None
    o_anki_durum = "Bosta (Bekleme)"
    tespit_edilen_alanlar = []
    tum_el_noktalari = []

    # 2. MediaPipe Analizi
    if MEDIAPIPE_HAZIR:
        kare_rgb = cv2.cvtColor(kare, cv2.COLOR_BGR2RGB)

        # ── 2a. Hands ──────────────────────────────────────────────────────────
        sonuclar = eller_dedektoru.process(kare_rgb)
        if sonuclar.multi_hand_landmarks:
            tum_el_noktalari = sonuclar.multi_hand_landmarks
            for el_noktalari in sonuclar.multi_hand_landmarks:
                bilek = el_noktalari.landmark[_mp_hands.HandLandmark.WRIST]
                bx, by = int(bilek.x * genislik), int(bilek.y * yukseklik)
                _mp_drawing.draw_landmarks(kare, el_noktalari, _mp_hands.HAND_CONNECTIONS)
                cv2.circle(kare, (bx, by), 8, (0, 0, 255), cv2.FILLED)
                for alan_adi, veri in tanimlanan_alanlar.items():
                    if point_in_polygon((bx, by), alan_poligonunu_al(veri)):
                        tespit_edilen_alanlar.append(alan_adi)

        # ── 2b. Grace Period & ID matching guncelle ──
        el_durumu_guncelle(ctx, tum_el_noktalari, kare_zamani, kare_suresi, yukseklik, genislik)

        # FSM Tracker guncellemesi (yalnızca eller dondurulmamışsa)
        if ctx.fsm_tracker is not None:
            aktif_eller = [el for el in [ctx.sol_el, ctx.sag_el] if el.aktif and not el.dondurulmus]
            if aktif_eller and sonuclar.multi_hand_landmarks:
                try:
                    ilk_el = el_landmarklarini_sozluge_cevir(sonuclar.multi_hand_landmarks[0])
                    ctx.fsm_tracker.update(ilk_el, (yukseklik, genislik))
                except Exception as e:
                    print(f"[UYARI] FSM güncellemesi başarısız: {e}")
            else:
                sol_grace = ctx.sol_el.grace_aktif
                sag_grace = ctx.sag_el.grace_aktif
                if not (sol_grace or sag_grace):
                    try:
                        ctx.fsm_tracker.update(None, (yukseklik, genislik))
                    except Exception as e:
                        print(f"[UYARI] FSM (None) güncellemesi başarısız: {e}")

        # ── 2c. Pose (her 2 karede bir) ─────────────────────────────────────
        if kare_sayaci % 2 == 0:
            try:
                pose_sonuc = pose_dedektoru.process(kare_rgb)
                if pose_sonuc.pose_landmarks:
                    lm = pose_sonuc.pose_landmarks.landmark
                    PL = _mp_pose.PoseLandmark

                    def lm_px(idx):
                        return (int(lm[idx].x * genislik), int(lm[idx].y * yukseklik))

                    sol_omuz   = lm_px(PL.LEFT_SHOULDER)
                    sol_dirsek = lm_px(PL.LEFT_ELBOW)
                    sol_bilek  = lm_px(PL.LEFT_WRIST)
                    sag_omuz   = lm_px(PL.RIGHT_SHOULDER)
                    sag_dirsek = lm_px(PL.RIGHT_ELBOW)
                    sag_bilek  = lm_px(PL.RIGHT_WRIST)
                    sol_kalca  = lm_px(PL.LEFT_HIP)

                    sol_aci = calculate_angle(sol_omuz, sol_dirsek, sol_bilek)
                    sag_aci = calculate_angle(sag_omuz, sag_dirsek, sag_bilek)
                    govde_acisi = calculate_trunk_angle(sol_kalca, sol_omuz)

                    ctx.ergo_state.is_pose_fresh = True
                    if not ctx.sol_el.dondurulmus:
                        ctx.ergo_state.dirsek_acisi_gecmisi_sol.append((sol_aci, True))
                        ctx.ergo_state.son_sol_aci = sol_aci
                    if not ctx.sag_el.dondurulmus:
                        ctx.ergo_state.dirsek_acisi_gecmisi_sag.append((sag_aci, True))
                        ctx.ergo_state.son_sag_aci = sag_aci
                    
                    if govde_acisi is not None:
                        ctx.ergo_state.govde_acisi_gecmisi.append((govde_acisi, True))
                        ctx.ergo_state.son_govde_acisi = govde_acisi

                    # İskelet çizimi
                    for (p1, p2) in [(sol_omuz, sol_dirsek), (sol_dirsek, sol_bilek),
                                     (sag_omuz, sag_dirsek), (sag_dirsek, sag_bilek)]:
                        cv2.line(kare, p1, p2, (180, 180, 180), 2)
                    for pt in [sol_omuz, sol_dirsek, sol_bilek,
                                sag_omuz, sag_dirsek, sag_bilek]:
                        cv2.circle(kare, pt, 5, (255, 255, 0), -1)

                    # Dirsek üzerine açı değeri
                    if not ctx.sol_el.dondurulmus:
                        cv2.putText(kare, f"{sol_aci:.0f}",
                                    (sol_dirsek[0] + 8, sol_dirsek[1]),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, aci_rengini_al(sol_aci), 2)
                    if not ctx.sag_el.dondurulmus:
                        cv2.putText(kare, f"{sag_aci:.0f}",
                                    (sag_dirsek[0] + 8, sag_dirsek[1]),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, aci_rengini_al(sag_aci), 2)
            except Exception as e:
                print(f"[UYARI] Pose analizi başarısız: {e}")
        else:
            ctx.ergo_state.is_pose_fresh = False
            if not ctx.sol_el.dondurulmus and ctx.ergo_state.son_sol_aci is not None:
                ctx.ergo_state.dirsek_acisi_gecmisi_sol.append((ctx.ergo_state.son_sol_aci, False))
            if not ctx.sag_el.dondurulmus and ctx.ergo_state.son_sag_aci is not None:
                ctx.ergo_state.dirsek_acisi_gecmisi_sag.append((ctx.ergo_state.son_sag_aci, False))
            if ctx.ergo_state.son_govde_acisi is not None:
                ctx.ergo_state.govde_acisi_gecmisi.append((ctx.ergo_state.son_govde_acisi, False))

    if tespit_edilen_alanlar:
        alet_alanlari = [a for a in tespit_edilen_alanlar
                         if tanimlanan_alanlar[a]["tip"] == "Alet Alanı"]
        elin_oldugu_aktif_alan = alet_alanlari[0] if alet_alanlari else tespit_edilen_alanlar[0]

    # --- Simülasyon (kütüphane yoksa) ----------------------------------------
    if not MEDIAPIPE_HAZIR and not elin_oldugu_aktif_alan:
        toplam = sum(ctx.hand_state.hareket_sureleri.values())
        alan_listesi = list(tanimlanan_alanlar.keys())
        if len(alan_listesi) > 0 and 2.0 < toplam <= 6.0:
            elin_oldugu_aktif_alan = alan_listesi[0]
        elif len(alan_listesi) > 1 and 9.0 < toplam <= 15.0:
            elin_oldugu_aktif_alan = alan_listesi[1]
    # -------------------------------------------------------------------------

    sol_grace = ctx.sol_el.grace_aktif
    sag_grace = ctx.sag_el.grace_aktif
    if sol_grace or sag_grace:
        o_anki_durum = "El Kaybi"
    else:
        # 4. Durum & Kronometre
        if elin_oldugu_aktif_alan:
            alan_tipi = tanimlanan_alanlar[elin_oldugu_aktif_alan]["tip"]
            if alan_tipi == "Calisma Alanı":
                o_anki_durum = "Montaj / Calisma"
                ctx.hand_state.son_bilinen_konum = "Calisma"
            else:
                ctx.hand_state.alan_kare_sayaclari[elin_oldugu_aktif_alan] += 1
                if ctx.hand_state.alan_kare_sayaclari[elin_oldugu_aktif_alan] >= 4:
                    o_anki_durum = "Malzeme Alma / Kavrama"
                    ctx.hand_state.son_bilinen_konum = "Alet"
                else:
                    o_anki_durum = "Kutulara Dogru Uzanma"
            ctx.hand_state.alan_sureleri[elin_oldugu_aktif_alan] += kare_suresi
        else:
            for alan in tanimlanan_alanlar:
                ctx.hand_state.alan_kare_sayaclari[alan] = 0
            if ctx.hand_state.son_bilinen_konum == "Alet":
                o_anki_durum = "Malzeme Tasimi (Masaya Dogru)"
            elif ctx.hand_state.son_bilinen_konum == "Calisma":
                o_anki_durum = "Kutulara Dogru Uzanma"
            else:
                o_anki_durum = "Bosta (Bekleme)"

        # 5. Süre Ekle
        if o_anki_durum in ctx.hand_state.hareket_sureleri:
            ctx.hand_state.hareket_sureleri[o_anki_durum] += kare_suresi

    # 6. Video süresi (sol üst)
    cv2.putText(kare,
                f"Video: {kare_sayaci * kare_suresi:.1f}s",
                (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 0, 255), 2)

    # 7. HUD Paneli ve FSM metrikleri
    fsm_metrikleri = None
    fsm_recipe = []
    fsm_recipe_idx = 0
    if ctx.fsm_tracker is not None:
        tracker = ctx.fsm_tracker
        fsm_metrikleri = tracker.get_metrics()
        fsm_recipe = getattr(tracker, "recipe", [])
        fsm_recipe_idx = getattr(tracker, "current_recipe_idx", 0)

        yeni_durum = tracker.state
        if ctx.fsm_eski_durum == "GRASP" and yeni_durum == "MOVE":
            ctx.fsm_events.append({"type": "GRASP", "timestamp": round(kare_zamani, 3)})
        elif ctx.fsm_eski_durum == "PLACE" and (yeni_durum == "REACH" or yeni_durum == "RETURNING_HOME" or yeni_durum == "IDLE"):
            ctx.fsm_events.append({"type": "PLACE", "timestamp": round(kare_zamani, 3)})
        ctx.fsm_eski_durum = yeni_durum

    # 8. Çevrim Takip (Bitiş Alanı)
    cycle_tracker_guncelle(ctx, tanimlanan_alanlar, tum_el_noktalari, kare_zamani, yukseklik, genislik)
    
    # 9. Flash & Banner Çizimi (Orijinal kareye çizelim, sonra resize edince de korunur)
    flash_ciz(kare, ctx, tanimlanan_alanlar)
    el_kayip_banner_ciz(kare, ctx, kare_zamani)

    # 10. Resizing and HUD Column Compiling
    if target_size is not None:
        target_w, target_h = target_size
        max_video_w = max(100, target_w - HUD_SERIT_GENISLIK)
        
        # En-boy oranını koruyarak video boyutunu hesapla
        v_ratio = genislik / yukseklik
        new_h = target_h
        new_w = int(target_h * v_ratio)
        if new_w > max_video_w:
            new_w = max_video_w
            new_h = int(max_video_w / v_ratio)
            
        # Video portion resize
        resized_video = cv2.resize(kare, (new_w, new_h))
        
        # Canvas oluştur (Video + HUD şeridi)
        canvas_w = new_w + HUD_SERIT_GENISLIK
        canvas_h = max(new_h, target_h)
        canvas = np.zeros((canvas_h, canvas_w, 3), dtype=np.uint8)
        
        # Videoyu sol tarafa yerleştir
        canvas[0:new_h, 0:new_w] = resized_video
        
        # HUD şeridini sağ tarafa çiz
        hud_serit_ciz(canvas, ctx, o_anki_durum, fsm_metrikleri, fsm_recipe, fsm_recipe_idx)
        return canvas
    else:
        # target_size yoksa orijinal kareye sağ şerit ekle
        canvas_w = genislik + HUD_SERIT_GENISLIK
        canvas = np.zeros((yukseklik, canvas_w, 3), dtype=np.uint8)
        canvas[0:yukseklik, 0:genislik] = kare
        hud_serit_ciz(canvas, ctx, o_anki_durum, fsm_metrikleri, fsm_recipe, fsm_recipe_idx)
        return canvas


# ── Rapor Kaydetme ────────────────────────────────────────────────────────────

def _xlsx_raporu_yaz(video_adi, alan_sureleri, hareket_sureleri, fsm_tracker, dirsek_acisi_gecmisi, govde_acisi_gecmisi, cycle_tracker=None, el_kayip_sure=0.0, el_kayip_olay=0):
    """openpyxl ile renkli, biçimlendirilmiş Excel (.xlsx) raporu oluşturur."""
    if not XLSX_HAZIR:
        return

    wb = openpyxl.Workbook()

    # ── Sayfa 1: Özet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Özet"

    baslik_doldur = PatternFill("solid", fgColor="1E3A5F")
    baslik_font   = Font(bold=True, color="FFFFFF", size=11)
    ust_baslik_doldur = PatternFill("solid", fgColor="2E86AB")
    ust_baslik_font   = Font(bold=True, color="FFFFFF", size=12)
    kenarlık = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    yesil = PatternFill("solid", fgColor="C8E6C9")
    sari  = PatternFill("solid", fgColor="FFF9C4")
    kirmizi = PatternFill("solid", fgColor="FFCDD2")

    def baslik_satir(ws, satir, metin, genis=4):
        ws.merge_cells(start_row=satir, start_column=1,
                       end_row=satir, end_column=genis)
        h = ws.cell(satir, 1, metin)
        h.fill = ust_baslik_doldur
        h.font = ust_baslik_font
        h.alignment = Alignment(horizontal="center")

    def ust_baslik(ws, satir, sutunlar):
        for c, metin in enumerate(sutunlar, 1):
            cell = ws.cell(satir, c, metin)
            cell.fill = baslik_doldur
            cell.font = baslik_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = kenarlık

    satir = 1
    ws.cell(satir, 1, f"Video: {video_adi}").font = Font(bold=True, size=13)
    satir += 2

    # Hareket Analizi
    baslik_satir(ws, satir, "HAREKET ANALİZİ")
    satir += 1
    ust_baslik(ws, satir, ["Hareket", "Süre (sn)", "Süre (TMU)", "Oran (%)"])
    satir += 1
    toplam_sure = sum(hareket_sureleri.values()) or 1
    for hareket, sure in hareket_sureleri.items():
        tmu = sure * 27.8
        oran = 100 * sure / toplam_sure
        for c, val in enumerate([hareket, f"{sure:.2f}", f"{tmu:.0f}", f"{oran:.1f}%"], 1):
            cell = ws.cell(satir, c, val)
            cell.border = kenarlık
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        satir += 1

    satir += 1
    # Ek Metrikler (El Kaybı)
    baslik_satir(ws, satir, "EL KAYBI METRİKLERİ")
    satir += 1
    ust_baslik(ws, satir, ["Metrik", "Değer", "", ""])
    satir += 1
    for c, val in enumerate(["Toplam El Kaybı Süresi (sn)", f"{el_kayip_sure:.2f}", "", ""], 1):
        cell = ws.cell(satir, c, val)
        cell.border = kenarlık
    satir += 1
    for c, val in enumerate(["El Kaybı Olay Sayısı", f"{el_kayip_olay}", "", ""], 1):
        cell = ws.cell(satir, c, val)
        cell.border = kenarlık
    satir += 2

    # Bölge / Alan Analizi
    baslik_satir(ws, satir, "BÖLGE / ALAN ANALİZİ")
    satir += 1
    ust_baslik(ws, satir, ["Alan Adı", "Süre (sn)", "", ""])
    satir += 1
    for alan, sure in alan_sureleri.items():
        for c, val in enumerate([alan, f"{sure:.2f}", "", ""], 1):
            cell = ws.cell(satir, c, val)
            cell.border = kenarlık
        satir += 1

    satir += 1
    # Dirsek Açısı Analizi
    if dirsek_acisi_gecmisi:
        baslik_satir(ws, satir, "DİRSEK AÇI ANALİZİ (Ergonomi)")
        satir += 1
        ust_baslik(ws, satir, ["Kol", "Ort Açı (°)", "Maks Açı (°)", "Yüksek Risk (%)", "Not"])
        satir += 1
        for taraf, aciler_tuples in dirsek_acisi_gecmisi.items():
            if not aciler_tuples:
                continue
            aciler = [a[0] for a in aciler_tuples]
            interpolated = any(not a[1] for a in aciler_tuples)
            ort  = float(np.mean(aciler))
            maks = float(np.max(aciler))
            risk = 100.0 * sum(1 for a in aciler if a >= ERGO_RISK_YUKSEK) / len(aciler)
            not_str = "Tahmini (Interpolated)" if interpolated else "Gerçek Zamanlı"
            for c, val in enumerate([taraf.capitalize(), f"{ort:.1f}", f"{maks:.1f}", f"{risk:.1f}%", not_str], 1):
                cell = ws.cell(satir, c, val)
                cell.border = kenarlık
                cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
            # Renk kodlama
            if maks >= ERGO_RISK_YUKSEK:
                ws.cell(satir, 3).fill = kirmizi
            elif maks >= ERGO_RISK_DIKKAT:
                ws.cell(satir, 3).fill = sari
            else:
                ws.cell(satir, 3).fill = yesil
            satir += 1

    if govde_acisi_gecmisi:
        baslik_satir(ws, satir, "GÖVDE AÇI ANALİZİ (Ergonomi)")
        satir += 1
        ust_baslik(ws, satir, ["Bölge", "Ort Açı (°)", "Maks Açı (°)", "Yüksek Risk (%)", "Not"])
        satir += 1
        
        aciler = [a[0] for a in govde_acisi_gecmisi]
        if aciler:
            interpolated = any(not a[1] for a in govde_acisi_gecmisi)
            ort  = float(np.mean(aciler))
            maks = float(np.max(aciler))
            risk = 100.0 * sum(1 for a in aciler if a >= GOVDE_RISK_YUKSEK) / len(aciler)
            not_str = "Tahmini (Interpolated)" if interpolated else "Gerçek Zamanlı"
            
            for c, val in enumerate(["Gövde", f"{ort:.1f}", f"{maks:.1f}", f"{risk:.1f}%", not_str], 1):
                cell = ws.cell(satir, c, val)
                cell.border = kenarlık
                cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
            
            if maks >= GOVDE_RISK_YUKSEK:
                ws.cell(satir, 3).fill = kirmizi
            elif maks >= GOVDE_RISK_DIKKAT:
                ws.cell(satir, 3).fill = sari
            else:
                ws.cell(satir, 3).fill = yesil
            satir += 1

    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 32
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 16

    # ── Sayfa 2: FSM Çevrimler ─────────────────────────────────────────────────
    if fsm_tracker is not None and getattr(fsm_tracker, "reported_cycles", None):
        ws2 = wb.create_sheet("FSM Çevrimler")
        ust_baslik(ws2, 1, ["Çevrim No", "İstasyon", "Durum Geçişleri",
                             "Süre (sn)", "TMU", "Sıra Doğru?"])
        for c in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws2.column_dimensions[c].width = 18
        for r, cyc in enumerate(fsm_tracker.reported_cycles, 2):
            degerler = [
                cyc["cycle_no"], cyc["station"], cyc["state_sequence"],
                cyc["duration_sec"], cyc["tmu"],
                "✓ Evet" if cyc["sequence_ok"] else "✗ Hayır"
            ]
            for c, val in enumerate(degerler, 1):
                cell = ws2.cell(r, c, val)
                cell.border = kenarlık
                cell.alignment = Alignment(horizontal="center")
            if not cyc["sequence_ok"]:
                for c in range(1, 7):
                    ws2.cell(r, c).fill = kirmizi

    # ── Sayfa 3: Çevrim Süresi (Cycle Time) ────────────────────────────────────
    if cycle_tracker is not None and (cycle_tracker.dongu_sureleri or cycle_tracker.atlanan_dongu_sureleri):
        ws3 = wb.create_sheet("Döngü & Çevrim")
        ust_baslik(ws3, 1, ["Döngü No", "Süre (sn)", "Ort.'dan Sapma", "Not"])
        for c in ['A', 'B', 'C', 'D']:
            ws3.column_dimensions[c].width = 18

        r = 2
        ort = cycle_tracker.cevrim_suresi
        
        for i, sure in enumerate(cycle_tracker.dongu_sureleri, 1):
            sapma = sure - ort
            sapma_str = f"{'+' if sapma>0 else ''}{sapma:.2f}s"
            
            for c, val in enumerate([i, f"{sure:.2f}s", sapma_str, "—"], 1):
                cell = ws3.cell(r, c, val)
                cell.border = kenarlık
                cell.alignment = Alignment(horizontal="center")
            r += 1

        for sure in cycle_tracker.atlanan_dongu_sureleri:
            for c, val in enumerate(["Atlandı", f"{sure:.2f}s", "—", "⚠️ Aykırı"], 1):
                cell = ws3.cell(r, c, val)
                cell.border = kenarlık
                cell.alignment = Alignment(horizontal="center")
            ws3.cell(r, 4).fill = sari
            r += 1

        r += 1
        ws3.cell(r, 1, "Toplam Ürün").font = Font(bold=True)
        ws3.cell(r, 2, cycle_tracker.urun_sayisi).font = Font(bold=True)
        r += 1
        ws3.cell(r, 1, "Ort. Çevrim").font = Font(bold=True)
        ws3.cell(r, 2, f"{ort:.2f}s").font = Font(bold=True)
        r += 1
        sapma_val = cycle_tracker.standart_sapma
        sapma_text = f"±{sapma_val:.2f}s" if sapma_val is not None else "N/A"
        ws3.cell(r, 1, "Std. Sapma").font = Font(bold=True)
        ws3.cell(r, 2, sapma_text).font = Font(bold=True)

    dosya_adi = f"{video_adi}_rapor.xlsx"
    wb.save(dosya_adi)
    print(f"[RAPOR] Excel kaydedildi: {dosya_adi}")


def excel_raporu_kaydet(video_adi, alan_sureleri, hareket_sureleri,
                        fsm_tracker=None, dirsek_acisi_gecmisi=None, govde_acisi_gecmisi=None, cycle_tracker=None, el_kayip_sure=0.0, el_kayip_olay=0):
    """CSV + XLSX (eğer openpyxl yüklüyse) raporları oluşturur."""

    # ── CSV (her zaman) ────────────────────────────────────────────────────────
    dosya_adi = f"{video_adi}_ergonomi_raporu.csv"
    with open(dosya_adi, mode='w', newline='', encoding='utf-8-sig') as f:
        yazici = csv.writer(f, delimiter=';')
        yazici.writerow(["Analiz Tipi", "Islem / Bölge Adı", "Süre / Değer"])

        yazici.writerow(["HAREKET ANALIZI", "", ""])
        for hareket, sure in hareket_sureleri.items():
            yazici.writerow(["", hareket, f"{sure:.2f}"])

        yazici.writerow(["EL KAYBI METRİKLERİ", "", ""])
        yazici.writerow(["", "Toplam El Kaybı Süresi (sn)", f"{el_kayip_sure:.2f}"])
        yazici.writerow(["", "El Kaybı Olay Sayısı", f"{el_kayip_olay}"])

        yazici.writerow(["BÖLGE/ALAN ANALIZI", "", ""])
        for alan, sure in alan_sureleri.items():
            yazici.writerow(["", alan, f"{sure:.2f}"])

        if dirsek_acisi_gecmisi:
            yazici.writerow(["DİRSEK AÇI ANALİZİ (Ergonomi)", "", ""])
            for taraf, aciler_tuples in dirsek_acisi_gecmisi.items():
                if aciler_tuples:
                    aciler = [a[0] for a in aciler_tuples]
                    interpolated = any(not a[1] for a in aciler_tuples)
                    ort  = np.mean(aciler)
                    maks = np.max(aciler)
                    risk = 100.0 * sum(1 for a in aciler if a >= ERGO_RISK_YUKSEK) / len(aciler)
                    yazici.writerow(["", f"{taraf.capitalize()} dirsek - Ort (°)",  f"{ort:.1f}"])
                    yazici.writerow(["", f"{taraf.capitalize()} dirsek - Maks (°)", f"{maks:.1f}"])
                    yazici.writerow(["", f"{taraf.capitalize()} dirsek - Risk (%)", f"{risk:.1f}"])
                    if interpolated:
                        yazici.writerow(["", f"{taraf.capitalize()} dirsek - Veri Kaynağı", "Tahmini (Interpolated)"])

        if govde_acisi_gecmisi:
            yazici.writerow(["GÖVDE AÇI ANALİZİ (Ergonomi)", "", ""])
            aciler = [a[0] for a in govde_acisi_gecmisi]
            if aciler:
                interpolated = any(not a[1] for a in govde_acisi_gecmisi)
                ort  = np.mean(aciler)
                maks = np.max(aciler)
                risk = 100.0 * sum(1 for a in aciler if a >= GOVDE_RISK_YUKSEK) / len(aciler)
                yazici.writerow(["", "Gövde - Ort (°)",  f"{ort:.1f}"])
                yazici.writerow(["", "Gövde - Maks (°)", f"{maks:.1f}"])
                yazici.writerow(["", "Gövde - Risk (%)", f"{risk:.1f}"])
                if interpolated:
                    yazici.writerow(["", "Gövde - Veri Kaynağı", "Tahmini (Interpolated)"])

    # FSM çevrim CSV
    if fsm_tracker is not None and getattr(fsm_tracker, "reported_cycles", None):
        fsm_csv = f"{video_adi}_fsm_raporu.csv"
        with open(fsm_csv, mode='w', newline='', encoding='utf-8-sig') as f:
            yazici = csv.writer(f, delimiter=';')
            yazici.writerow(["Cevrim No", "Istasyon", "Durum Gecisleri",
                             "Sure (sn)", "TMU", "Sira Dogru Mu"])
            for c in fsm_tracker.reported_cycles:
                yazici.writerow([
                    c["cycle_no"], c["station"], c["state_sequence"],
                    c["duration_sec"], c["tmu"], c["sequence_ok"]
                ])

    # ── XLSX (openpyxl varsa) ──────────────────────────────────────────────────
    if XLSX_HAZIR:
        try:
            _xlsx_raporu_yaz(video_adi, alan_sureleri, hareket_sureleri, fsm_tracker, dirsek_acisi_gecmisi, govde_acisi_gecmisi, cycle_tracker, el_kayip_sure, el_kayip_olay)
        except Exception as e:
            print(f"[UYARI] Excel (.xlsx) dosyası yazılamadı: {e}")


def generate_llm_report(video_adi, ctx):
    """Token izlemeli, anti-halüsinasyon promptlu ve bağımlılık korumalı metin analizi."""
    import os
    import json
    import re
    from datetime import datetime
    from config_manager import GEMINI_MODEL_NAME

    # Kütüphanenin yüklü olmama ihtimalini yakalıyoruz (Import Hatası Koruması)
    try:
        import google.generativeai as genai
    except ImportError:
        print("[LLM RAPORU] google-generativeai kütüphanesi kurulu değil. AI raporu atlandı.")
        return "Yapay Zeka Raporu şu an oluşturulamadı (Gerekli kütüphane yüklü değil)."

    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        print("[LLM RAPORU] API anahtarı eksik, rapor oluşturulamadı.")
        return "Yapay Zeka Raporu şu an oluşturulamadı (API anahtarı bulunamadı)."

    ct = getattr(ctx, 'cycle_tracker', None)
    if not ct:
        print("[LLM RAPORU] HATA: cycle_tracker bulunamadı.")
        return "Veri modeli eksik."

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(GEMINI_MODEL_NAME)
        
        toplam_sure = sum(ctx.hand_state.hareket_sureleri.values()) or 1.0
        durum_dagilimi = {k: f"%{(v/toplam_sure)*100:.1f}" for k, v in ctx.hand_state.hareket_sureleri.items()}

        prompt = f"""
        Aşağıda endüstriyel bir montaj hattında yapılan zaman etüdü ölçümleri yer almaktadır:
        - Analiz Edilen Video: {video_adi}
        - Toplam Ürün Sayısı: {ct.urun_sayisi}
        - Ortalama Çevrim Süresi: {ct.cevrim_suresi:.2f} saniye (Standart Sapma: {ct.standart_sapma if ct.standart_sapma else 0:.2f}s)
        - El Kayıp Toplam Süresi: {ctx.el_kayip_toplam_sure:.2f} saniye (Olay Sayısı: {ctx.el_kayip_olay_sayisi})
        - Durum Süre Dağılımları: {json.dumps(durum_dagilimi, ensure_ascii=False)}
        
        Bu verileri analiz et ve bir endüstri mühendisi perspektifinden özet rapor yaz.
        
        KRİTİK KURALLAR:
        1. Yalnızca yukarıda verilen sayısal verilere sadık kal.
        2. Kanıtlanamayan veya spekülatif neden-sonuç ilişkileri kurma (Örn: 'İşçi yorulduğu için yavaşladı' deme. Sadece 'Çevrim süreleri arasında yüksek sapma görüldü' tespiti yap).
        3. Raporu kısa, net ve aksiyona yönelik 3-4 maddede Türkçe olarak yaz.
        """
        
        response = model.generate_content(
            prompt,
            request_options={"timeout": 15.0}
        )
        
        if hasattr(response, 'usage_metadata') and response.usage_metadata:
            meta = response.usage_metadata
            print(f"[LLM RAPORU] Model: {GEMINI_MODEL_NAME} | Input: {meta.prompt_token_count} | Output: {meta.candidates_token_count}")
            
        report_text = response.text + "\n\n⚠️ Bu rapor yapay zeka tarafından otomatik üretilmiştir. Karar almadan önce mühendis onayı gereklidir."
        
        safe_video_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', os.path.basename(video_adi))
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"{safe_video_name}_{timestamp}_ai_analiz.txt")
        
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report_text)
            
        print(f"[LLM RAPORU] Rapor başarıyla yazıldı: {report_path}")
        return report_text
        
    except Exception as e:
        print(f"[LLM RAPORU] Rapor oluşturulamadı (Graceful Degradation): {e}")
        return "Yapay Zeka Raporu şu an oluşturulamadı (Bağlantı/Sistem Hatası)."